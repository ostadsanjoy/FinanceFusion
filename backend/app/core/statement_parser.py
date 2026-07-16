import io
from datetime import datetime, date

import openpyxl
import pdfplumber

from app.core.categories import CATEGORIES, guess_category

REQUIRED_COLUMNS = ["DATE", "AMOUNT", "RECEIVER", "PURPOSE"]
OPTIONAL_COLUMNS = ["REMARKS"]

# Accepts common header spelling/naming variants and maps them to the
# canonical name used everywhere else in this module.
HEADER_ALIASES = {
    "RECIEVER": "RECEIVER",  # the sheet's actual (misspelled) header
    "RECEIVER": "RECEIVER",
    "DATE": "DATE",
    "AMOUNT": "AMOUNT",
    "PURPOSE": "PURPOSE",
    "CATEGORY": "PURPOSE",
    "REMARKS": "REMARKS",
    "REMARK": "REMARKS",
    "NOTES": "REMARKS",
}


def _find_columns(header_row):
    """Maps column name -> index if this row looks like a transactions
    header row (has all of DATE/AMOUNT/RECEIVER/PURPOSE). Used to both
    detect which sheet/table is the real data and to locate columns
    regardless of their order.

    The same page can contain a second, smaller "SPENDS SPLITS" summary
    table (category + total) sitting to the right of the main table —
    pdfplumber sometimes merges both into one detected table. That summary
    table also has a column literally called "Amount", so we only keep the
    *first* occurrence of each column name (the real one, further left)
    and ignore any later duplicate.
    """
    idx = {}
    for i, cell in enumerate(header_row or []):
        if cell is None:
            continue
        key = HEADER_ALIASES.get(str(cell).strip().upper())
        if key:
            idx.setdefault(key, i)
    if all(col in idx for col in REQUIRED_COLUMNS):
        return idx
    return None


def _parse_date(value, fallback_year):
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())

    text = str(value).strip()
    if not text or text.lower() == "none":
        return None

    # A row like "04/04-11/04" is a summed multi-day entry — use the
    # range's start date.
    if "-" in text:
        text = text.split("-")[0].strip()

    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue

    # The sheet usually shows dates as "DD/MM" with no year — the year is
    # implied by which monthly tab the row lives on, which we lose once
    # exported to PDF. Fall back to the year supplied at import time.
    try:
        parsed = datetime.strptime(text, "%d/%m")
        return parsed.replace(year=fallback_year)
    except ValueError:
        return None


def _parse_amount(value):
    if isinstance(value, (int, float)):
        return float(value)
    text = (
        str(value)
        .replace(",", "")
        .replace("₹", "")
        .replace("Rs.", "")
        .replace("Rs", "")
        .strip()
    )
    try:
        return float(text)
    except ValueError:
        return None


def _normalize_category(purpose_text, receiver_text):
    text = str(purpose_text or "").strip().lower()
    if text in CATEGORIES:
        return text
    return guess_category(f"{receiver_text} {purpose_text}")


def _rows_to_transactions(rows, fallback_year):
    transactions = []
    skipped = 0

    for row in rows:
        idx = row["idx"]
        cells = row["cells"]

        def cell(name):
            i = idx.get(name)
            return cells[i] if i is not None and i < len(cells) else None

        raw_date = cell("DATE")
        if raw_date is None or str(raw_date).strip() == "" or str(raw_date).strip().lower() == "none":
            continue

        parsed_date = _parse_date(raw_date, fallback_year)
        parsed_amount = _parse_amount(cell("AMOUNT"))

        if parsed_date is None or parsed_amount is None:
            skipped += 1
            continue

        raw_receiver = cell("RECEIVER")
        raw_purpose = cell("PURPOSE")
        raw_remarks = cell("REMARKS")

        receiver = str(raw_receiver).strip() if raw_receiver else "Expense"
        remarks = str(raw_remarks).strip() if raw_remarks else ""
        description = f"{receiver} — {remarks}" if remarks else receiver

        transactions.append(
            {
                "date": parsed_date,
                "amount": parsed_amount,
                "description": description,
                "category": _normalize_category(raw_purpose, receiver),
                "source_type": "CASH",
                "is_expense": True,
            }
        )

    return transactions, skipped


def parse_xlsx(content: bytes, fallback_year: int):
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    rows = []

    for ws in wb.worksheets:
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        idx = _find_columns(header)
        if not idx:
            continue  # not a transactions sheet (e.g. a yearly overview tab)

        for r in ws.iter_rows(min_row=2, values_only=True):
            rows.append({"idx": idx, "cells": list(r)})

    return _rows_to_transactions(rows, fallback_year)


def parse_pdf(content: bytes, fallback_year: int):
    rows = []
    idx = None

    with pdfplumber.open(io.BytesIO(content)) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if not table:
                continue
            for raw_row in table:
                if idx is None:
                    candidate = _find_columns(raw_row)
                    if candidate:
                        idx = candidate
                    continue
                rows.append({"idx": idx, "cells": raw_row})

    if idx is None:
        return [], 0

    return _rows_to_transactions(rows, fallback_year)


def parse_statement(filename: str, content: bytes, fallback_year: int):
    lower = (filename or "").lower()
    if lower.endswith(".xlsx") or lower.endswith(".xlsm"):
        return parse_xlsx(content, fallback_year)
    if lower.endswith(".pdf"):
        return parse_pdf(content, fallback_year)
    raise ValueError("Unsupported file type — please upload a .xlsx or .pdf file")