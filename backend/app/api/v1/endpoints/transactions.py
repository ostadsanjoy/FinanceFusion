import io
from datetime import datetime
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from google.cloud.firestore_v1.base_query import FieldFilter
from openpyxl import Workbook
from openpyxl.chart import DoughnutChart, Reference
from openpyxl.styles import Font, PatternFill

from app.database import get_db
from app.schemas import transaction as schemas
from app.api.v1.endpoints.auth import get_current_user
from app.core.categories import CATEGORIES
from app.core.statement_parser import parse_statement

router = APIRouter()

TRANSACTIONS = "transactions"


@router.post("/", response_model=schemas.Transaction)
def create_transaction(
    transaction: schemas.TransactionCreate,
    db=Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    data = transaction.model_dump()
    data["date"] = data["date"] or datetime.utcnow()
    data["user_id"] = current_user["id"]

    _, doc_ref = db.collection(TRANSACTIONS).add(data)
    data["id"] = doc_ref.id
    return data


@router.get("/", response_model=List[schemas.Transaction])
def read_transactions(
    skip: int = 0,
    limit: int = 1000,
    db=Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    query = (
        db.collection(TRANSACTIONS)
        .where(filter=FieldFilter("user_id", "==", current_user["id"]))
        .offset(skip)
        .limit(limit)
    )
    results = []
    for doc in query.stream():
        data = doc.to_dict()
        data["id"] = doc.id
        results.append(data)
    return results


@router.post("/import")
async def import_transactions(
    file: UploadFile = File(...),
    year: Optional[int] = Form(None),
    db=Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    content = await file.read()
    fallback_year = year or datetime.utcnow().year

    try:
        parsed, skipped = parse_statement(file.filename, content, fallback_year)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    if not parsed:
        raise HTTPException(
            status_code=400,
            detail="No recognizable transaction rows found. Make sure the file has "
            "DATE, AMOUNT, RECEIVER, and PURPOSE columns.",
        )

    collection_ref = db.collection(TRANSACTIONS)

    # Firestore batched writes are capped at 500 operations, so chunk large imports.
    CHUNK = 400
    for start in range(0, len(parsed), CHUNK):
        batch = db.batch()
        for tx in parsed[start : start + CHUNK]:
            doc_ref = collection_ref.document()
            batch.set(doc_ref, {**tx, "user_id": current_user["id"]})
        batch.commit()

    return {"imported": len(parsed), "skipped": skipped}


@router.get("/export")
def export_transactions(db=Depends(get_db), current_user: dict = Depends(get_current_user)):
    docs = (
        db.collection(TRANSACTIONS)
        .where(filter=FieldFilter("user_id", "==", current_user["id"]))
        .stream()
    )
    txs = [doc.to_dict() for doc in docs]
    txs.sort(key=lambda t: t.get("date") or datetime.min)

    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    headers = ["DATE", "AMOUNT", "RECEIVER", "PURPOSE", "REMARKS"]
    header_fill = PatternFill(start_color="FFD9EAD3", end_color="FFD9EAD3", fill_type="solid")
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True)
        c.fill = header_fill

    for i, tx in enumerate(txs, start=2):
        raw_date = tx.get("date")
        description = tx.get("description", "") or ""
        if " — " in description:
            receiver, remarks = description.split(" — ", 1)
        else:
            receiver, remarks = description, ""

        ws.cell(
            row=i,
            column=1,
            value=raw_date.strftime("%d/%m/%Y") if hasattr(raw_date, "strftime") else str(raw_date),
        )
        ws.cell(row=i, column=2, value=tx.get("amount"))
        ws.cell(row=i, column=3, value=receiver)
        ws.cell(row=i, column=4, value=tx.get("category"))
        ws.cell(row=i, column=5, value=remarks)
    summary_col = 7
    ws.cell(row=1, column=summary_col, value="SPENDS SPLITS").font = Font(bold=True)
    ws.cell(row=1, column=summary_col + 1, value="AMOUNT").font = Font(bold=True)

    totals = {cat: 0.0 for cat in CATEGORIES}
    for tx in txs:
        cat = (tx.get("category") or "misc").lower()
        totals[cat] = totals.get(cat, 0.0) + (tx.get("amount") or 0)

    row = 2
    for cat in CATEGORIES:
        ws.cell(row=row, column=summary_col, value=cat.capitalize())
        ws.cell(row=row, column=summary_col + 1, value=round(totals.get(cat, 0.0), 2))
        row += 1
    ws.cell(row=row, column=summary_col, value="Total spend").font = Font(bold=True)
    ws.cell(
        row=row, column=summary_col + 1, value=round(sum(totals.values()), 2)
    ).font = Font(bold=True)

    chart = DoughnutChart()
    chart.title = "Spends Split"
    data_ref = Reference(ws, min_col=summary_col + 1, min_row=2, max_row=row - 1)
    cats_ref = Reference(ws, min_col=summary_col, min_row=2, max_row=row - 1)
    chart.add_data(data_ref, titles_from_data=False)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, "J2")

    for col_letter, width in zip("ABCDE", [12, 10, 24, 16, 30]):
        ws.column_dimensions[col_letter].width = width
    for col_letter, width in zip("GH", [16, 12]):
        ws.column_dimensions[col_letter].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"spends_export_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.put("/{transaction_id}", response_model=schemas.Transaction)
def update_transaction(
    transaction_id: str,
    transaction_update: schemas.TransactionCreate,
    db=Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    doc_ref = db.collection(TRANSACTIONS).document(transaction_id)
    snap = doc_ref.get()
    if not snap.exists or snap.to_dict().get("user_id") != current_user["id"]:
        raise HTTPException(status_code=404, detail="Transaction not found")

    update_data = transaction_update.model_dump()
    if update_data["date"] is None:
        update_data["date"] = snap.to_dict().get("date")

    doc_ref.update(update_data)
    data = doc_ref.get().to_dict()
    data["id"] = doc_ref.id
    return data


@router.delete("/{transaction_id}")
def delete_transaction(
    transaction_id: str,
    db=Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    doc_ref = db.collection(TRANSACTIONS).document(transaction_id)
    snap = doc_ref.get()
    if not snap.exists or snap.to_dict().get("user_id") != current_user["id"]:
        raise HTTPException(status_code=404, detail="Transaction not found")

    doc_ref.delete()
    return {"message": "Transaction deleted"}